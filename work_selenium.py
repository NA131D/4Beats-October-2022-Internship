import openpyxl
import time
from selenium import webdriver
from datetime import datetime as date

def scrapfunc(e):
    global long,short
    driver=webdriver.Edge()
    # driver.minimize_window()
    driver.get("https://www.google.com/")
    sbox=driver.find_element("xpath",".//*[@title='সার্চ করুন']")
    sbox.send_keys(e)
    time.sleep(5)
    sug=driver.find_elements("xpath","//ul[@role='listbox']/li/div/div/div/span")
    for row in sug:
        x.append(row.text)
    xx = list(filter(None, x))    
    driver.close()   
    long=max(xx,key=len)
    short=min(xx,key=len)
    x.clear()
    xx.clear()
    return long,short
day=date.today().strftime("%A")
print(str(day))
kw=[]
x=[]
path = "E:\PY\getdata.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.sheetnames
ws=wb[day]
for cell in ws['C']:
    if cell.value is not None:
             kw.append((cell.value))       
i=0        
while i<len(kw):
    scrapfunc(kw[i])
    cv=str(3+i)
    ws['D'+cv]=long
    ws['E'+cv]=short
    wb.save('E:\PY\getdata.xlsx')
    i=i+1





        